#!/usr/bin/env python3
"""Generate a 14-week calendar sheet in an Excel workbook."""

import argparse
import sys
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


GRAY_FILL = PatternFill(patternType="solid", fgColor="F0F0F0")
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def prev_sunday(d: date) -> date:
    """Return d if it's a Sunday, else the preceding Sunday."""
    return d - timedelta(days=d.weekday() + 1 if d.weekday() != 6 else 0)


def cell_label(d: date, prev: date | None) -> str:
    """Return 'Mon D' at a month boundary, else just 'D'."""
    if prev is None or d.month != prev.month:
        return f"{MONTHS[d.month - 1]} {d.day}"
    return str(d.day)


def _core_months(start: date, weeks: int = 14) -> list[date]:
    """Months with 14+ days in the range (trims fringe months at either end)."""
    end = start + timedelta(weeks=weeks) - timedelta(days=1)
    result = []
    cur = start.replace(day=1)
    while cur <= end:
        next_month = date(cur.year + (cur.month == 12), cur.month % 12 + 1, 1)
        days_in_range = (min(next_month - timedelta(days=1), end) -
                         max(cur, start)).days + 1
        if days_in_range >= 14:
            result.append(cur)
        cur = next_month
    return result or [start.replace(day=1)]


def sheet_title(start: date, weeks: int = 14) -> str:
    months = _core_months(start, weeks)
    if len(months) == 1:
        return months[0].strftime("%B %Y")
    first, last = months[0], months[-1]
    if first.year == last.year:
        return f"{first.strftime('%B')} - {last.strftime('%B')} {first.year}"
    return f"{first.strftime('%B %Y')} - {last.strftime('%B %Y')}"


def tab_name(start: date, weeks: int = 14) -> str:
    months = _core_months(start, weeks)
    if len(months) == 1:
        return months[0].strftime("%b %Y")
    first, last = months[0], months[-1]
    if first.year == last.year:
        return f"{first.strftime('%b')}-{last.strftime('%b')} {first.year}"
    return f"{first.strftime('%b %Y')} - {last.strftime('%b %Y')}"


def build_sheet(wb: openpyxl.Workbook, start: date, weeks: int = 14):
    name = tab_name(start, weeks)
    if name in wb.sheetnames:
        print(f"Sheet '{name}' already exists — overwriting.", file=sys.stderr)
        del wb[name]

    ws = wb.create_sheet(title=name)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="top")

    # Row 1: merged title
    ws.merge_cells("A1:H1")
    ws["A1"] = sheet_title(start, weeks)
    ws["A1"].font = bold
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30.75

    # Row 2: day headers
    for col, h in enumerate(["Notes", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = bold
        cell.alignment = center
    ws.row_dimensions[2].height = 15

    ws.column_dimensions["A"].width = 7.13
    ws.column_dimensions["B"].width = 8.38

    all_days = [start + timedelta(days=i) for i in range(weeks * 7)]

    for week in range(weeks):
        date_row = 3 + week * 2
        empty_row = date_row + 1
        ws.row_dimensions[date_row].height = 17.25
        ws.row_dimensions[empty_row].height = 17.25
        ws.merge_cells(f"A{date_row}:A{empty_row}")
        ws[f"A{date_row}"].alignment = center

        for day_offset in range(7):
            idx = week * 7 + day_offset
            d = all_days[idx]
            prev = all_days[idx - 1] if idx > 0 else None
            cell = ws.cell(row=date_row, column=2 + day_offset, value=cell_label(d, prev))
            cell.alignment = left
            if day_offset in (0, 6):
                cell.fill = GRAY_FILL

    return ws


def parse_date(s: str) -> date:
    try:
        return date.fromisoformat(s)
    except ValueError:
        print(f"Error: invalid date '{s}'. Use YYYY-MM-DD format.", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Add a 14-week calendar sheet to an Excel workbook.",
        epilog="Example: gen_calendar.py 2026-04-05",
    )
    parser.add_argument(
        "start_date",
        help="Start date (YYYY-MM-DD). Snapped to the preceding Sunday.",
    )
    parser.add_argument(
        "--file", "-f",
        default="Calendar Templates.xlsx",
        help="Excel file path (default: 'Calendar Templates.xlsx')",
    )
    parser.add_argument(
        "--weeks", "-w",
        type=int,
        default=14,
        help="Number of weeks (default: 14)",
    )
    args = parser.parse_args()

    start = prev_sunday(parse_date(args.start_date))

    try:
        wb = openpyxl.load_workbook(args.file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    build_sheet(wb, start, args.weeks)
    wb.save(args.file)

    tab = tab_name(start, args.weeks)
    end = start + timedelta(weeks=args.weeks) - timedelta(days=1)
    print(f"Added sheet '{tab}' ({sheet_title(start, args.weeks)}) to {args.file}")
    print(f"  {start} – {end} ({args.weeks} weeks)")


if __name__ == "__main__":
    main()
